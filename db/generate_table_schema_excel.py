"""
Script to generate Excel file with Risks and Users table columns
Run: python generate_table_schema_excel.py
Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create a new workbook
wb = openpyxl.Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Define Risks table columns
risks_columns = [
    {'Column Name': 'RiskId', 'Data Type': 'UNIQUEIDENTIFIER', 'Nullable': 'NOT NULL', 'Description': 'Primary Key - Unique identifier for the risk'},
    {'Column Name': 'RiskNo', 'Data Type': 'NVARCHAR', 'Nullable': 'NULL', 'Description': 'Risk number (e.g., O001, M002)'},
    {'Column Name': 'DepartmentId', 'Data Type': 'UNIQUEIDENTIFIER', 'Nullable': 'NOT NULL', 'Description': 'Foreign Key to Departments table'},
    {'Column Name': 'Name', 'Data Type': 'NVARCHAR(MAX)', 'Nullable': 'NULL', 'Description': 'Risk name (nullable)'},
    {'Column Name': 'Description', 'Data Type': 'NVARCHAR(MAX)', 'Nullable': 'NOT NULL', 'Description': 'Risk description'},
    {'Column Name': 'CategoryId', 'Data Type': 'NVARCHAR(200)', 'Nullable': 'NULL', 'Description': 'Category name (stored as string)'},
    {'Column Name': 'Identification', 'Data Type': 'NVARCHAR(50)', 'Nullable': 'NULL', 'Description': 'Inherent risk or Residual risk'},
    {'Column Name': 'ExistingControlInPlace', 'Data Type': 'NVARCHAR(1000)', 'Nullable': 'NULL', 'Description': 'Existing controls description'},
    {'Column Name': 'PlanOfAction', 'Data Type': 'NVARCHAR(1000)', 'Nullable': 'NULL', 'Description': 'Plan of action description'},
    {'Column Name': 'Impact', 'Data Type': 'NVARCHAR', 'Nullable': 'NOT NULL', 'Description': 'Severe, Significant, Moderate, Minor, Negligible'},
    {'Column Name': 'Likelihood', 'Data Type': 'NVARCHAR', 'Nullable': 'NOT NULL', 'Description': 'Very likely, Likely, Possible, Unlikely, Very Unlikely'},
    {'Column Name': 'Status', 'Data Type': 'NVARCHAR', 'Nullable': 'NOT NULL', 'Description': 'Raised, Rejected, Open, Closed, In Progress, New, Existing, Downgraded, Upgraded, Eliminated'},
    {'Column Name': 'OwnerId', 'Data Type': 'UNIQUEIDENTIFIER', 'Nullable': 'NOT NULL', 'Description': 'Foreign Key to Owners table'},
    {'Column Name': 'RejectionReason', 'Data Type': 'NVARCHAR(1000)', 'Nullable': 'NULL', 'Description': 'Reason for rejection (if status is Rejected)'},
    {'Column Name': 'CreatedByUserId', 'Data Type': 'UNIQUEIDENTIFIER', 'Nullable': 'NULL', 'Description': 'Foreign Key to Users table - who created the risk'},
    {'Column Name': 'CreatedAtUtc', 'Data Type': 'DATETIME', 'Nullable': 'NOT NULL', 'Description': 'Creation timestamp (UTC)'},
    {'Column Name': 'UpdatedAtUtc', 'Data Type': 'DATETIME', 'Nullable': 'NOT NULL', 'Description': 'Last update timestamp (UTC)'},
]

# Define Users table columns
users_columns = [
    {'Column Name': 'UserId', 'Data Type': 'UNIQUEIDENTIFIER', 'Nullable': 'NOT NULL', 'Description': 'Primary Key - Unique identifier for the user'},
    {'Column Name': 'Name', 'Data Type': 'NVARCHAR', 'Nullable': 'NOT NULL', 'Description': 'User full name'},
    {'Column Name': 'Email', 'Data Type': 'NVARCHAR(256)', 'Nullable': 'NULL', 'Description': 'User email address'},
    {'Column Name': 'Role', 'Data Type': 'NVARCHAR', 'Nullable': 'NOT NULL', 'Description': 'user, manager, admin, or unit_head'},
    {'Column Name': 'DepartmentId', 'Data Type': 'UNIQUEIDENTIFIER', 'Nullable': 'NULL', 'Description': 'Foreign Key to Departments table'},
    {'Column Name': 'EmployeeId', 'Data Type': 'NVARCHAR(128)', 'Nullable': 'NULL', 'Description': 'Employee ID (6 digits + @kauveryhospital.com)'},
    {'Column Name': 'Unit', 'Data Type': 'NVARCHAR(50)', 'Nullable': 'NULL', 'Description': 'Unit code (e.g., KCN, KTN, KCH)'},
    {'Column Name': 'IsUnitHead', 'Data Type': 'BIT', 'Nullable': 'NOT NULL', 'Description': 'Boolean flag indicating if user is a unit head (default: 0)'},
]

# Create Risks sheet
ws_risks = wb.create_sheet('Risks')
ws_risks.append(['Column Name', 'Data Type', 'Nullable', 'Description'])

# Style header row
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws_risks[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add Risks data
for col in risks_columns:
    ws_risks.append([col['Column Name'], col['Data Type'], col['Nullable'], col['Description']])

# Auto-adjust column widths
ws_risks.column_dimensions['A'].width = 25
ws_risks.column_dimensions['B'].width = 25
ws_risks.column_dimensions['C'].width = 15
ws_risks.column_dimensions['D'].width = 50

# Create Users sheet
ws_users = wb.create_sheet('Users')
ws_users.append(['Column Name', 'Data Type', 'Nullable', 'Description'])

# Style header row
for cell in ws_users[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add Users data
for col in users_columns:
    ws_users.append([col['Column Name'], col['Data Type'], col['Nullable'], col['Description']])

# Auto-adjust column widths
ws_users.column_dimensions['A'].width = 25
ws_users.column_dimensions['B'].width = 25
ws_users.column_dimensions['C'].width = 15
ws_users.column_dimensions['D'].width = 50

# Save the workbook
output_file = 'database_schema.xlsx'
wb.save(output_file)
print(f'Excel file created successfully: {output_file}')
print(f'Risks table: {len(risks_columns)} columns')
print(f'Users table: {len(users_columns)} columns')

